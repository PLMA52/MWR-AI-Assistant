"""
Age Breakout from Dan's US_IncomeByZipDemographics file
Reads ZIP-level age data from IncomebyzipCodeReport tab (columns BR-BY),
aggregates to county level, calculates percentages, writes to Neo4j and CSV.

Age Groups (Dan's 8 groups):
  0-9, 10-19, 20-29, 30-39, 40-49, 50-59, 60-69, 70+
"""

import os
import csv
from neo4j import GraphDatabase
from openpyxl import load_workbook

# ============================================================
# CONFIGURATION
# ============================================================
NEO4J_URI = os.getenv("NEO4J_URI", "neo4j+s://551c1b37.databases.neo4j.io")
NEO4J_USERNAME = os.getenv("NEO4J_USERNAME", "neo4j")
NEO4J_PASSWORD = os.getenv("NEO4J_PASSWORD", "RYA3u0QAGRSbpdjJpB4r5-yWvAVPW4rdioPy00lRx3g")

# Dan's file - use the light version with just the needed tabs
EXCEL_FILE = os.path.expanduser("~/OneDrive - SODEXO/MWR_Automation_Data/US_IncomeByZipDemographics_LightVersion.xlsx")
# If light version doesn't exist, try the original
EXCEL_FILE_ALT = os.path.expanduser("~/OneDrive - SODEXO/MWR_Automation_Data/US_IncomeByZipDemographics (Table format with FIPS).xlsb")

OUTPUT_CSV = os.path.expanduser("~/OneDrive - SODEXO/MWR_Automation_Data/Census_Age_Breakout_Jan2026.csv")

# State name to abbreviation mapping
STATE_NAME_TO_ABBR = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "District of Columbia": "DC", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI",
    "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME",
    "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN",
    "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE",
    "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM",
    "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH",
    "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX",
    "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
    "Puerto Rico": "PR", "Virgin Islands": "VI", "Guam": "GU",
}

# Column indices in IncomeByZipCodeReport (1-based from openpyxl)
COL_STATE = 9      # I: state_name
COL_COUNTY = 10    # J: county_name
COL_AGE_0_9 = 70   # BR: age_Total_0_to_9
COL_AGE_10_19 = 71  # BS: age_Total_10_to_19
COL_AGE_20_29 = 72  # BT: age_Total_20_to_29
COL_AGE_30_39 = 73  # BU: age_Total_30_to_39
COL_AGE_40_49 = 74  # BV: age_Total_40_to_49
COL_AGE_50_59 = 75  # BW: age_Total_50_to_59
COL_AGE_60_69 = 76  # BX: age_Total_60_to_69
COL_AGE_70_PLUS = 77  # BY: age_Total_70_plus


def safe_int(val):
    """Safely convert to int, defaulting to 0"""
    try:
        return int(val) if val is not None and val != "" else 0
    except (ValueError, TypeError):
        return 0


def read_excel_data():
    """Read ZIP-level age data from Dan's Excel file"""
    print("=" * 60)
    print("STEP 1: READ AGE DATA FROM EXCEL")
    print("=" * 60)
    
    file_path = EXCEL_FILE
    if not os.path.exists(file_path):
        file_path = EXCEL_FILE_ALT
        if not os.path.exists(file_path):
            print(f"   ERROR: File not found!")
            print(f"   Tried: {EXCEL_FILE}")
            print(f"   Tried: {EXCEL_FILE_ALT}")
            print(f"   Please copy the light version to your OneDrive MWR_Automation_Data folder.")
            return None
    
    print(f"   Reading: {file_path}")
    print(f"   This may take a moment for large files...")
    
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb["IncomeByZipCodeReport"]
    
    # Aggregate to county level
    county_data = {}  # key: (county, state_abbr) -> age sums
    row_count = 0
    skipped = 0
    
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_count += 1
        
        # Get state and county
        state_name = row[COL_STATE - 1].value  # 0-indexed
        county_name = row[COL_COUNTY - 1].value
        
        if not state_name or not county_name:
            skipped += 1
            continue
        
        # Convert state name to abbreviation
        state_abbr = STATE_NAME_TO_ABBR.get(state_name)
        if not state_abbr:
            skipped += 1
            continue
        
        # Get age values
        age_0_9 = safe_int(row[COL_AGE_0_9 - 1].value)
        age_10_19 = safe_int(row[COL_AGE_10_19 - 1].value)
        age_20_29 = safe_int(row[COL_AGE_20_29 - 1].value)
        age_30_39 = safe_int(row[COL_AGE_30_39 - 1].value)
        age_40_49 = safe_int(row[COL_AGE_40_49 - 1].value)
        age_50_59 = safe_int(row[COL_AGE_50_59 - 1].value)
        age_60_69 = safe_int(row[COL_AGE_60_69 - 1].value)
        age_70_plus = safe_int(row[COL_AGE_70_PLUS - 1].value)
        
        # Handle multi-county entries (e.g., "Abbeville; Greenwood; McCormick")
        # Split on semicolons and assign to EACH county listed
        if ";" in str(county_name):
            counties = [c.strip() for c in county_name.split(";") if c.strip()]
        else:
            counties = [county_name.strip()]
        
        # Distribute age data to each county
        # For multi-county ZIPs, we assign the full count to each county
        # (since we're computing percentages, not absolute counts, this is fine)
        for single_county in counties:
            key = (single_county, state_abbr)
        
            if key not in county_data:
                county_data[key] = {
                    "age_0_to_9": 0, "age_10_to_19": 0, "age_20_to_29": 0,
                    "age_30_to_39": 0, "age_40_to_49": 0, "age_50_to_59": 0,
                    "age_60_to_69": 0, "age_70_plus": 0
                }
        
            county_data[key]["age_0_to_9"] += age_0_9
            county_data[key]["age_10_to_19"] += age_10_19
            county_data[key]["age_20_to_29"] += age_20_29
            county_data[key]["age_30_to_39"] += age_30_39
            county_data[key]["age_40_to_49"] += age_40_49
            county_data[key]["age_50_to_59"] += age_50_59
            county_data[key]["age_60_to_69"] += age_60_69
            county_data[key]["age_70_plus"] += age_70_plus
        
        if row_count % 5000 == 0:
            print(f"      Processed {row_count} ZIPs...")
    
    wb.close()
    
    print(f"\n   Total ZIPs processed: {row_count}")
    print(f"   ZIPs skipped (missing state/county): {skipped}")
    print(f"   Counties with data: {len(county_data)}")
    
    return county_data


def calculate_percentages(county_data):
    """Calculate age group percentages for each county"""
    print("\n" + "=" * 60)
    print("STEP 2: CALCULATE PERCENTAGES")
    print("=" * 60)
    
    county_results = {}
    
    for (county, state), data in county_data.items():
        total = (data["age_0_to_9"] + data["age_10_to_19"] + data["age_20_to_29"] +
                 data["age_30_to_39"] + data["age_40_to_49"] + data["age_50_to_59"] +
                 data["age_60_to_69"] + data["age_70_plus"])
        
        if total == 0:
            continue
        
        county_results[(county, state)] = {
            "total_age_pop": total,
            "age_0_to_9": data["age_0_to_9"],
            "age_10_to_19": data["age_10_to_19"],
            "age_20_to_29": data["age_20_to_29"],
            "age_30_to_39": data["age_30_to_39"],
            "age_40_to_49": data["age_40_to_49"],
            "age_50_to_59": data["age_50_to_59"],
            "age_60_to_69": data["age_60_to_69"],
            "age_70_plus": data["age_70_plus"],
            "pct_0_to_9": round(data["age_0_to_9"] / total * 100, 1),
            "pct_10_to_19": round(data["age_10_to_19"] / total * 100, 1),
            "pct_20_to_29": round(data["age_20_to_29"] / total * 100, 1),
            "pct_30_to_39": round(data["age_30_to_39"] / total * 100, 1),
            "pct_40_to_49": round(data["age_40_to_49"] / total * 100, 1),
            "pct_50_to_59": round(data["age_50_to_59"] / total * 100, 1),
            "pct_60_to_69": round(data["age_60_to_69"] / total * 100, 1),
            "pct_70_plus": round(data["age_70_plus"] / total * 100, 1),
        }
    
    print(f"   Counties with valid percentages: {len(county_results)}")
    
    # Validation
    print("\n   VALIDATION - Sample counties:")
    test_counties = [
        ("Jefferson", "AL"),
        ("Los Angeles", "CA"),
        ("San Francisco", "CA"),
        ("Boulder", "CO"),
        ("Montgomery", "MD"),
    ]
    for county, state in test_counties:
        if (county, state) in county_results:
            d = county_results[(county, state)]
            print(f"\n   {county}, {state} (pop: {d['total_age_pop']:,}):")
            print(f"      0-9: {d['pct_0_to_9']}%  |  10-19: {d['pct_10_to_19']}%  |  20-29: {d['pct_20_to_29']}%  |  30-39: {d['pct_30_to_39']}%")
            print(f"      40-49: {d['pct_40_to_49']}%  |  50-59: {d['pct_50_to_59']}%  |  60-69: {d['pct_60_to_69']}%  |  70+: {d['pct_70_plus']}%")
            pct_sum = (d['pct_0_to_9'] + d['pct_10_to_19'] + d['pct_20_to_29'] +
                       d['pct_30_to_39'] + d['pct_40_to_49'] + d['pct_50_to_59'] +
                       d['pct_60_to_69'] + d['pct_70_plus'])
            print(f"      Sum of %: {pct_sum}%")
    
    return county_results


def export_csv(county_results):
    """Export county-level age data to CSV for Power BI"""
    print("\n" + "=" * 60)
    print("STEP 3: EXPORT CSV FOR POWER BI")
    print("=" * 60)
    
    print(f"   Exporting to: {OUTPUT_CSV}")
    
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            "County", "State", "Total_Age_Population",
            "Age_0_to_9", "Age_10_to_19", "Age_20_to_29", "Age_30_to_39",
            "Age_40_to_49", "Age_50_to_59", "Age_60_to_69", "Age_70_Plus",
            "Pct_0_to_9", "Pct_10_to_19", "Pct_20_to_29", "Pct_30_to_39",
            "Pct_40_to_49", "Pct_50_to_59", "Pct_60_to_69", "Pct_70_Plus"
        ])
        
        for (county, state), data in sorted(county_results.items()):
            writer.writerow([
                county, state, data["total_age_pop"],
                data["age_0_to_9"], data["age_10_to_19"], data["age_20_to_29"], data["age_30_to_39"],
                data["age_40_to_49"], data["age_50_to_59"], data["age_60_to_69"], data["age_70_plus"],
                data["pct_0_to_9"], data["pct_10_to_19"], data["pct_20_to_29"], data["pct_30_to_39"],
                data["pct_40_to_49"], data["pct_50_to_59"], data["pct_60_to_69"], data["pct_70_plus"]
            ])
    
    print(f"   Exported {len(county_results)} counties")


def write_to_neo4j(county_results):
    """Write age percentages to Neo4j ZipCode nodes"""
    print("\n" + "=" * 60)
    print("STEP 4: WRITE TO NEO4J")
    print("=" * 60)
    
    driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USERNAME, NEO4J_PASSWORD))
    updated_total = 0
    not_found = 0
    
    try:
        with driver.session() as session:
            for (county, state), data in county_results.items():
                result = session.run('''
                    MATCH (z:ZipCode)
                    WHERE z.county = $county AND z.state = $state
                    SET z.pct_age_0_to_9 = $p1,
                        z.pct_age_10_to_19 = $p2,
                        z.pct_age_20_to_29 = $p3,
                        z.pct_age_30_to_39 = $p4,
                        z.pct_age_40_to_49 = $p5,
                        z.pct_age_50_to_59 = $p6,
                        z.pct_age_60_to_69 = $p7,
                        z.pct_age_70_plus = $p8
                    RETURN count(z) as cnt
                ''', county=county, state=state,
                   p1=data["pct_0_to_9"], p2=data["pct_10_to_19"],
                   p3=data["pct_20_to_29"], p4=data["pct_30_to_39"],
                   p5=data["pct_40_to_49"], p6=data["pct_50_to_59"],
                   p7=data["pct_60_to_69"], p8=data["pct_70_plus"])
                
                cnt = result.single()["cnt"]
                if cnt > 0:
                    updated_total += cnt
                else:
                    not_found += 1
                
                if updated_total % 5000 == 0 and updated_total > 0:
                    print(f"      Updated {updated_total} ZIPs...")
    finally:
        driver.close()
    
    print(f"\n   Updated {updated_total} ZIPs with age breakout data")
    print(f"   Counties not found in Neo4j: {not_found}")
    
    # Verify
    print(f"\n   VERIFICATION:")
    driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USERNAME, NEO4J_PASSWORD))
    try:
        with driver.session() as session:
            result = session.run('''
                MATCH (z:ZipCode)
                WHERE z.county = "San Francisco" AND z.state = "CA"
                RETURN DISTINCT z.county, z.state,
                    z.pct_age_0_to_9, z.pct_age_10_to_19, z.pct_age_20_to_29,
                    z.pct_age_30_to_39, z.pct_age_40_to_49, z.pct_age_50_to_59,
                    z.pct_age_60_to_69, z.pct_age_70_plus
                LIMIT 1
            ''')
            record = result.single()
            if record:
                print(f"   San Francisco, CA:")
                print(f"      0-9: {record['z.pct_age_0_to_9']}%  |  10-19: {record['z.pct_age_10_to_19']}%")
                print(f"      20-29: {record['z.pct_age_20_to_29']}%  |  30-39: {record['z.pct_age_30_to_39']}%")
                print(f"      40-49: {record['z.pct_age_40_to_49']}%  |  50-59: {record['z.pct_age_50_to_59']}%")
                print(f"      60-69: {record['z.pct_age_60_to_69']}%  |  70+: {record['z.pct_age_70_plus']}%")
            
            # Count coverage
            result2 = session.run('''
                MATCH (z:ZipCode) WHERE z.pct_age_0_to_9 IS NOT NULL
                RETURN count(z) as cnt
            ''')
            cnt = result2.single()["cnt"]
            print(f"   Total ZIPs with age data: {cnt}")
    finally:
        driver.close()


def main():
    # Step 1: Read Excel
    county_data = read_excel_data()
    if county_data is None:
        return
    
    # Step 2: Calculate percentages
    county_results = calculate_percentages(county_data)
    
    # Step 3: Export CSV
    export_csv(county_results)
    
    # Step 4: Write to Neo4j
    write_to_neo4j(county_results)
    
    print("\n" + "=" * 60)
    print("DONE! Age breakout data loaded.")
    print("Next steps:")
    print("  1. Add Census_Age_Breakout_Jan2026.csv to Power BI Inform tab")
    print("  2. Update AI Assistant prompt with age properties")
    print("  3. Push updates to GitHub")
    print("=" * 60)


if __name__ == "__main__":
    main()
